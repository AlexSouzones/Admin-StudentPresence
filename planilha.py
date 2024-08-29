import pandas as pd
import openpyxl
import shutil


def planilha_insert(
    dados: list[tuple[str]],
    nome_arquivo_copia: str,
    modelo_arquivo: str = "planilha/model/model.xlsx",
):
    nome_arquivo_copia = f"planilha/planilhas geradas/{nome_arquivo_copia.upper()}.xlsx"
    shutil.copyfile(modelo_arquivo, nome_arquivo_copia)

    workbook = openpyxl.load_workbook(nome_arquivo_copia)
    sheet = workbook.active

    for row in sheet.iter_rows(min_row=2, max_col=1, max_row=sheet.max_row):
        if not row[0].value:
            primeira_linha_vazia = row[0].row
            break
    else:
        primeira_linha_vazia = sheet.max_row + 1

    for item in dados:
        linha = [item[0], item[1], item[2], item[3], item[4], item[5], item[7].upper()]
        for col_num, value in enumerate(linha, start=1):
            sheet.cell(row=primeira_linha_vazia, column=col_num, value=value)
        primeira_linha_vazia += 1

    workbook.save(nome_arquivo_copia)


class PlanilhaManager:
    def __init__(self, file: str) -> None:
        self.file = file
        self.disciplina: str = None
        self.polos: list[str] = []
        self.info: list[dict[str, list[dict[str]]]] = []

    def polo_content(self, polo_argument: str) -> bool:
        for dictionary in self.info:
            if polo_argument == dictionary["Polo"]:
                return True
        return False

    def insert_into_polo(self, polo: str, info_data: dict[str]):
        for dictionary in self.info:
            if polo == dictionary["Polo"]:
                dictionary["Alunos"].append(info_data)

    def get_disciplina(self) -> str:
        try:
            self.disciplina = self.info[0]["Disciplina"]
            return self.disciplina
        except:
            return False

    def loading(self):
        df = pd.read_excel(self.file)
        for index, row in df.iterrows():
            if not self.disciplina:
                self.disciplina = row["DISCIPLINA"]
            polo = row["POLO EAD"]
            cod = row["COD POLO"]
            info_aluno = {
                "matricula": f"{row['RA']}",
                "nome": f"{row['ALUNO']}",
                "Curso": f"{row['CURSO']}",
                "Disciplina": f"{self.disciplina}",
            }
            if not self.polo_content(polo):
                self.info.append(
                    {
                        "COD POLO": f"{cod}",
                        "Polo": f"{polo}",
                        "Alunos": [info_aluno],
                    }
                )
            else:
                self.insert_into_polo(polo, info_aluno)


if __name__ == "__main__":
    ...
